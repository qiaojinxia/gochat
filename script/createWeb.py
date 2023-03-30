import pandas as pd
from openpyxl import load_workbook
if __name__ == '__main__':
    # 加载Excel文件
    book = load_workbook('/data/goods list.xlsx')
    writer = pd.ExcelWriter('example', engine='openpyxl')

    # 遍历所有工作表
    for sheet_name in book.sheetnames:
        # 读取工作表数据
        sheet = book[sheet_name]
        data = [cell.value for row in sheet for cell in row]

        # 将数据整理成DataFrame格式
        rows = sheet.max_row
        cols = sheet.max_column
        df_data = [data[x:x+cols] for x in range(0, len(data), cols)]
        df = pd.DataFrame(df_data)
        # 渲染HTML并添加样式
        df.to_html(writer, sheet_name, index=False)
        writer.sheets[sheet_name].column_dimensions.group("A", cols, hidden=True)
    # 保存HTML文件
    writer.save()